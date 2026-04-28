"""
Analyse ChatGPT broadband-prompt JSON/JSONL files and produce a summary Excel.

Reads every .json / .jsonl file in a folder, then for each record:
  - Strips URLs and inline markdown citations from the `response` text
  - Counts mentions of each canonical broadband brand
  - Extracts citation domains from the `sources` array (falling back to regex)

Outputs a single Excel file with five tabs:
  1. Cleaned Responses        - one row per response, with source file
  2. Brand Counts by File     - wide table, rows = files, cols = brands
  3. Citation Counts by File  - long table, (file, domain, count)
  4. Brand Totals             - overall brand -> count across all files
  5. Citation Totals          - overall domain -> count across all files

Usage:
    python analyse_chatgpt_responses.py <input_folder> [output_xlsx]

If no args given, defaults to current dir and ./chatgpt_analysis.xlsx
"""

from __future__ import annotations

import json
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path
from urllib.parse import urlparse

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# CONFIG: edit these to tweak the analysis
# ---------------------------------------------------------------------------

BRANDS: dict[str, list[str]] = {
    "Sky":             ["Sky", "Sky Broadband"],
    "Virgin Media":    ["Virgin Media", "Virgin"],
    "BT":              ["BT", "BT Broadband", "BT Full Fibre"],
    "TalkTalk":        ["TalkTalk"],
    "Plusnet":         ["Plusnet"],
    "Vodafone":        ["Vodafone"],
    "EE":              ["EE", "EE Broadband"],
    "Hyperoptic":      ["Hyperoptic"],
    "Zen Internet":    ["Zen Internet", "Zen"],
    "Community Fibre": ["Community Fibre"],
    "NOW Broadband":   ["NOW Broadband", "NOW"],
    "YouFibre":        ["YouFibre"],
    "Openreach":       ["Openreach"],
    "Gigaclear":       ["Gigaclear"],
    "Three":           ["Three"],
    "Cuckoo":          ["Cuckoo"],
}

CASE_SENSITIVE_BRANDS: set[str] = {"EE", "BT", "NOW Broadband", "Three"}


# ---------------------------------------------------------------------------
# Regex patterns
# ---------------------------------------------------------------------------

# Inline markdown citation:  ([example.com](https://example.com/path))
# Also matches without the outer parens: [example.com](https://example.com)
INLINE_CITATION_RE = re.compile(
    r"\(?\[[^\]]*\]\(https?://[^)]+\)\)?"
)
# Any bare URL left over after the inline-citation pass
BARE_URL_RE = re.compile(r"https?://\S+")
# Tidy up doubled whitespace left behind
WHITESPACE_RE = re.compile(r"[ \t]{2,}")
NEWLINE_RUNS_RE = re.compile(r"\n{3,}")


# ---------------------------------------------------------------------------
# Brand-matching: pre-compile one regex per canonical brand
# ---------------------------------------------------------------------------

def _build_brand_patterns() -> dict[str, re.Pattern]:
    """Compile one regex per canonical brand, longest variants first."""
    patterns: dict[str, re.Pattern] = {}
    for canonical, variants in BRANDS.items():
        # Sort longest first so 'Sky Broadband' wins over 'Sky' when both could
        # match at the same position (re.findall is non-overlapping).
        sorted_variants = sorted(variants, key=len, reverse=True)
        escaped = [re.escape(v) for v in sorted_variants]
        # Word boundaries stop 'BT' matching inside 'obtain', 'EE' inside 'see', etc.
        joined = r"\b(?:" + "|".join(escaped) + r")\b"
        flags = 0 if canonical in CASE_SENSITIVE_BRANDS else re.IGNORECASE
        patterns[canonical] = re.compile(joined, flags)
    return patterns


BRAND_PATTERNS = _build_brand_patterns()


# ---------------------------------------------------------------------------
# Cleaning
# ---------------------------------------------------------------------------

def clean_response(text: str) -> str:
    """Strip inline markdown citations and bare URLs from a response."""
    if not text:
        return ""
    text = INLINE_CITATION_RE.sub("", text)
    text = BARE_URL_RE.sub("", text)
    text = WHITESPACE_RE.sub(" ", text)
    text = NEWLINE_RUNS_RE.sub("\n\n", text)
    # Tidy stray punctuation left after stripping ' (... )' style citations
    text = re.sub(r"\(\s*\)", "", text)
    text = re.sub(r"\s+([,.;:])", r"\1", text)
    return text.strip()


# ---------------------------------------------------------------------------
# Counting
# ---------------------------------------------------------------------------

def count_brands(clean_text: str) -> dict[str, int]:
    """Count canonical-brand mentions in the cleaned text."""
    return {
        canonical: len(pattern.findall(clean_text))
        for canonical, pattern in BRAND_PATTERNS.items()
    }


def domain_from_url(url: str) -> str | None:
    """Return a normalised domain (lowercase, no leading 'www.') or None."""
    if not url:
        return None
    try:
        host = urlparse(url).netloc.lower()
    except Exception:
        return None
    if not host:
        return None
    return host.removeprefix("www.")


def extract_citation_domains(record: dict, raw_response: str) -> list[str]:
    """Return a list of domains cited (with repetition) for one record.

    Prefers the structured `sources` list; falls back to regex on the raw
    response if that field is missing.
    """
    domains: list[str] = []

    sources = record.get("sources")
    if isinstance(sources, list) and sources:
        for src in sources:
            url = src.get("url") if isinstance(src, dict) else None
            d = domain_from_url(url) if url else None
            if d:
                domains.append(d)
        return domains

    # Fallback: pull URLs out of the raw response text
    for url in BARE_URL_RE.findall(raw_response or ""):
        d = domain_from_url(url.rstrip(").,;"))
        if d:
            domains.append(d)
    return domains


# ---------------------------------------------------------------------------
# File loading
# ---------------------------------------------------------------------------

def load_records(path: Path) -> list[dict]:
    """Load records from a .json or .jsonl file. Returns a list of dicts."""
    text = path.read_text(encoding="utf-8")
    suffix = path.suffix.lower()

    # JSONL: one object per line
    if suffix == ".jsonl":
        out = []
        for i, line in enumerate(text.splitlines(), start=1):
            line = line.strip()
            if not line:
                continue
            try:
                out.append(json.loads(line))
            except json.JSONDecodeError as e:
                print(f"  ! {path.name} line {i}: {e}", file=sys.stderr)
        return out

    # JSON: either a single object or an array of objects
    try:
        data = json.loads(text)
    except json.JSONDecodeError:
        # Some "json" files are actually jsonl; try line-by-line as a fallback
        out = []
        for line in text.splitlines():
            line = line.strip()
            if line:
                try:
                    out.append(json.loads(line))
                except json.JSONDecodeError:
                    pass
        return out

    if isinstance(data, list):
        return [d for d in data if isinstance(d, dict)]
    if isinstance(data, dict):
        return [data]
    return []


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def analyse_folder(folder: Path) -> dict[str, pd.DataFrame]:
    """Walk the folder and build the four output DataFrames."""
    files = sorted(
        [p for p in folder.iterdir() if p.suffix.lower() in {".json", ".jsonl"}]
    )
    if not files:
        raise SystemExit(f"No .json or .jsonl files found in {folder}")

    response_rows: list[dict] = []
    brand_counts_by_file: dict[str, Counter] = defaultdict(Counter)
    domain_counts_by_file: dict[str, Counter] = defaultdict(Counter)

    for path in files:
        records = load_records(path)
        print(f"  - {path.name}: {len(records)} records")

        for rec in records:
            raw_response = rec.get("response", "") or ""
            cleaned = clean_response(raw_response)

            # Tab 1 row
            response_rows.append({
                "source_file": path.name,
                "prompt": rec.get("prompt", ""),
                "run_number": rec.get("run_number"),
                "timestamp": rec.get("timestamp"),
                "model": rec.get("model"),
                "status": rec.get("status"),
                "response_clean": cleaned,
                "response_char_count": len(cleaned),
            })

            # Tab 2 totals
            for brand, n in count_brands(cleaned).items():
                if n:
                    brand_counts_by_file[path.name][brand] += n

            # Tab 3 totals
            for domain in extract_citation_domains(rec, raw_response):
                domain_counts_by_file[path.name][domain] += 1

    # ---- Tab 1: cleaned responses ----
    df_responses = pd.DataFrame(response_rows)

    # ---- Tab 2: brand counts by file (wide) ----
    brand_cols = list(BRANDS.keys())
    brand_rows = []
    for fname in [p.name for p in files]:
        row = {"source_file": fname}
        row.update({b: brand_counts_by_file[fname].get(b, 0) for b in brand_cols})
        row["TOTAL"] = sum(row[b] for b in brand_cols)
        brand_rows.append(row)
    df_brands = pd.DataFrame(brand_rows, columns=["source_file", *brand_cols, "TOTAL"])
    # Append column totals as a final row
    totals_row = {"source_file": "TOTAL"}
    for b in brand_cols:
        totals_row[b] = int(df_brands[b].sum())
    totals_row["TOTAL"] = int(df_brands["TOTAL"].sum())
    df_brands = pd.concat([df_brands, pd.DataFrame([totals_row])], ignore_index=True)

    # ---- Tab 3: citation counts by file (long) ----
    citation_rows = []
    for fname, counter in domain_counts_by_file.items():
        for domain, n in counter.items():
            citation_rows.append({
                "source_file": fname,
                "domain": domain,
                "count": n,
            })
    df_citations = pd.DataFrame(citation_rows).sort_values(
        ["source_file", "count", "domain"], ascending=[True, False, True]
    ).reset_index(drop=True)

    # ---- Tab 4: brand totals (across all files) ----
    overall_brands = Counter()
    for c in brand_counts_by_file.values():
        overall_brands.update(c)
    df_brand_totals = pd.DataFrame(
        [{"brand": b, "total_mentions": overall_brands.get(b, 0)} for b in brand_cols]
    ).sort_values("total_mentions", ascending=False).reset_index(drop=True)

    # ---- Tab 5: citation totals (across all files) ----
    overall_domains = Counter()
    for c in domain_counts_by_file.values():
        overall_domains.update(c)
    df_domain_totals = pd.DataFrame(
        [{"domain": d, "total_citations": n} for d, n in overall_domains.items()]
    ).sort_values("total_citations", ascending=False).reset_index(drop=True)

    return {
        "Cleaned Responses": df_responses,
        "Brand Counts by File": df_brands,
        "Citation Counts by File": df_citations,
        "Brand Totals": df_brand_totals,
        "Citation Totals": df_domain_totals,
    }


# ---------------------------------------------------------------------------
# Excel writing + light formatting
# ---------------------------------------------------------------------------

def write_excel(sheets: dict[str, pd.DataFrame], out_path: Path) -> None:
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Pretty up: bold header, freeze top row, sensible widths
    wb = load_workbook(out_path)
    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    header_fill = PatternFill("solid", start_color="305496")
    body_font = Font(name="Arial")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.freeze_panes = "A2"

        # Style header row
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")

        # Body font + column widths
        for col_idx, col in enumerate(ws.columns, start=1):
            letter = get_column_letter(col_idx)
            max_len = len(str(ws.cell(row=1, column=col_idx).value or ""))
            for cell in col:
                if cell.row > 1:
                    cell.font = body_font
                val = cell.value
                if val is not None:
                    L = len(str(val))
                    if L > max_len:
                        max_len = L
            # Cap so a giant cleaned-response cell doesn't make a 5000-wide column
            ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 60)

        # On the responses tab, wrap the big text column so it's readable
        if sheet_name == "Cleaned Responses":
            headers = [c.value for c in ws[1]]
            if "response_clean" in headers:
                col_idx = headers.index("response_clean") + 1
                letter = get_column_letter(col_idx)
                ws.column_dimensions[letter].width = 80
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")

        # Bold the TOTAL row/column on the brand counts tab
        if sheet_name == "Brand Counts by File":
            last_row = ws.max_row
            last_col = ws.max_column
            for cell in ws[last_row]:
                cell.font = Font(bold=True, name="Arial")
            for row in ws.iter_rows(min_row=2, min_col=last_col, max_col=last_col):
                for cell in row:
                    cell.font = Font(bold=True, name="Arial")

    wb.save(out_path)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    if len(sys.argv) >= 2:
        folder = Path(sys.argv[1]).expanduser().resolve()
    else:
        folder = Path.cwd()

    if len(sys.argv) >= 3:
        out_path = Path(sys.argv[2]).expanduser().resolve()
    else:
        out_path = folder / "chatgpt_analysis.xlsx"

    if not folder.is_dir():
        raise SystemExit(f"Not a directory: {folder}")

    print(f"Scanning: {folder}")
    sheets = analyse_folder(folder)
    write_excel(sheets, out_path)
    print(f"\nWrote: {out_path}")
    for name, df in sheets.items():
        print(f"  {name}: {len(df):,} rows")


if __name__ == "__main__":
    main()
